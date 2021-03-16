from pathlib import Path

import openpyxl
from django.contrib.auth.models import User
from django.db import models

# Create your models here.
class File(models.Model):
    file_name=models.CharField(max_length=20,verbose_name='نام فایل')
    file=models.FileField(upload_to='uploads/')
    user=models.ForeignKey(User,on_delete=models.CASCADE,related_name='file')

    def __str__(self):
        return self.file_name

    def save(self,*args,**kwargs):
        super().save(*args, **kwargs)
        xlsx_file = Path(self.file.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        for row in sheet.iter_rows(max_row=sheet.max_row):
            user = SeminarCustomer.objects.create(first_name=row[0].value, last_name=row[1].value, phone=row[2].value, )

            user.file.add(self)

        return

class SeminarCustomer(models.Model):
    first_name=models.CharField(max_length=20,verbose_name='نام')
    last_name=models.CharField(max_length=20,verbose_name='نام خانوادگی')
    phone=models.IntegerField(default=0)
    file=models.ManyToManyField(File)
    come=models.BooleanField(default=False)
    backup_user=models.ForeignKey(User,on_delete=models.CASCADE,default=1,related_name='customer')


    def __str__(self):
        return self.first_name+" "+self.last_name


# class BackupUser(models.Model):
#     user=models.OneToOneField(User,on_delete=models.CASCADE)
#     seminar_customer=models.ForeignKey(SeminarCustomer,on_delete=models.CASCADE)

